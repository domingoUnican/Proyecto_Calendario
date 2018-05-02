# -*- coding: utf-8 -*-
import openpyxl
import mysql.connector
import collections
import unicodedata
#import unidecode

##### CAMBIAR A LA INFO PERSONALIZADA EN CADA EQUIPO
#conectarse a la BD ###CAMBIAR
cnx = mysql.connector.connect(user='root', password='itsasecret',
                              host='127.0.0.1',
                              database='scheduleDB')


#fijar la constante de nombre y asegurar que estamos conectados a la BD correcta
DB_NAME='scheduleDB' ###CAMBIAR
filepath = "C:\\Users\\Manuel\\Desktop\\Proyecto_Calendario-master\\Modulo BD\\PythonCode\\test.xlsx"   ###CAMBIAR
############################





#Correspondencia entre periodos de tiempo y sus correspondientes casillas en el horario
days_times={"Lunes1": "B2", "Lunes2": "B3", "Lunes3":"B4", "Lunes4":"B5", "Lunes5":"B6", "Lunes10":"B7", "Lunes6":"B9", "Lunes7":"B10", "Lunes8":"B11", "Lunes9":"B12",
            "Martes1":"C2", "Martes2":"C3", "Martes3":"C4", "Martes4":"C5", "Martes5":"C6", "Martes10":"C7", "Martes6":"C9", "Martes7":"C10", "Martes8":"C11", "Martes9":"C12",
            "Miercoles1":"D2", "Miercoles2":"D3", "Miercoles3":"D4", "Miercoles4":"D5", "Miercoles5":"D6", "Miercoles10":"D7", "Miercoles6":"D9", "Miercoles7":"D10", 
            "Miercoles8":"D11", "Miercoles9":"D12",
            "Jueves1":"E2", "Jueves2":"E3", "Jueves3":"E4", "Jueves4":"E5", "Jueves5":"E6", "Jueves10":"E7", "Jueves6":"E9", "Jueves7":"E10", "Jueves8":"E11", "Jueves9":"E12",
            "Viernes1":"F2", "Viernes2":"F3", "Viernes3":"F4", "Viernes4":"F5", "Viernes5":"F6", "Viernes10":"F7", "Viernes6":"F9", "Viernes7":"F10", "Viernes8":"F11", "Viernes9":"F12"
            }

#Lista de grupos a buscar junto al nombre de la hoja de excel que deberá generarse
pares_grupo_hoja=[("'Class1_INFORMATICA_PLACEHOLDER';","info_1_c2"),
       ("'Class1_MATEMATICAS_PLACEHOLDER';","mate_1_c2"),
       ("'Class1_FISICA_PLACEHOLDER';","fisica_1_c2"),
       ("'Class1_DOBLE_PLACEHOLDER';","doble_1_c2"),
       ("'Class2_INFORMATICA_PLACEHOLDER';","info_2_c2"),
       ("'Class2_MATEMATICAS_PLACEHOLDER';","mate_2_c2"),
       ("'Class2_FISICA_PLACEHOLDER';","fisica_2_c2"),
       ("'Class2_DOBLE_PLACEHOLDER';","doble_2_c2"),
       ("'Class3_INFORMATICA_PLACEHOLDER';","info_3_c2"),
       ("'Class3_MATEMATICAS_PLACEHOLDER';","mate_3_c2"),
       ("'Class3_FISICA_PLACEHOLDER';","fisica_3_c2"),
       ("'Class3_DOBLE_PLACEHOLDER';","doble_3_c2")]


cnx.database=DB_NAME

#obtener el cursor para operar con la BD
cursor=cnx.cursor()

#definimos las queries
query1=("SELECT events_cuatrimestre2_1718.eventgroup, events_cuatrimestre2_1718.id, events_cuatrimestre2_1718.time, res_aulas.name, res_grAlumnos.name FROM  events_cuatrimestre2_1718"
    "INNER JOIN events_need_resources_c2_1718 enr_aulas ON enr_aulas.event= events_cuatrimestre2_1718.id"
    "INNER JOIN resources res_aulas ON enr_aulas.resource=res_aulas.id"
    "INNER JOIN events_need_resources_c2_1718 enr_grAlumnos ON enr_grAlumnos.event= events_cuatrimestre2_1718.id"
    "INNER JOIN resources res_grAlumnos ON enr_grAlumnos.resource=res_grAlumnos.id"
    
    "WHERE events_cuatrimestre2_1718.eventgroup='%s'"   #introducir aquí código de asignatura 
        "AND res_grAlumnos.resourcetype='Class'"
        "AND res_aulas.resourcetype='Room';")
query2=('''
SELECT events_cuatrimestre2_1718.eventgroup, events_cuatrimestre2_1718.id, events_cuatrimestre2_1718.time, res_aulas.name, res_aulas.color, res_gr_alumnos.id FROM  events_cuatrimestre2_1718
    INNER JOIN events_need_resources_c2_1718 enr_aulas ON enr_aulas.event= events_cuatrimestre2_1718.id
    INNER JOIN resources res_aulas ON enr_aulas.resource=res_aulas.id
    INNER JOIN events_need_resources_c2_1718 enr_gr_alumnos ON enr_gr_alumnos.event= events_cuatrimestre2_1718.id
    INNER JOIN resources res_gr_alumnos ON enr_gr_alumnos.resource=res_gr_alumnos.id
    
    WHERE res_gr_alumnos.id='Class1_INFORMATICA_PLACEHOLDER'
        AND res_gr_alumnos.resourcetype='Class'
        AND res_aulas.resourcetype='Room'
        
        ORDER BY events_cuatrimestre2_1718.time
        ;
''')

query3=("SELECT events_cuatrimestre2_1718.eventgroup, events_cuatrimestre2_1718.id, events_cuatrimestre2_1718.time, res_aulas.name, res_aulas.color, res_gr_alumnos.id, ev_gr.name FROM  events_cuatrimestre2_1718"
    "INNER JOIN events_need_resources_c2_1718 enr_aulas ON enr_aulas.event= events_cuatrimestre2_1718.id"
    "INNER JOIN resources res_aulas ON enr_aulas.resource=res_aulas.id"
    "INNER JOIN events_need_resources_c2_1718 enr_gr_alumnos ON enr_gr_alumnos.event= events_cuatrimestre2_1718.id"
    "INNER JOIN resources res_gr_alumnos ON enr_gr_alumnos.resource=res_gr_alumnos.id"
    "INNER JOIN eventgroups ev_gr ON events_cuatrimestre2_1718.eventgroup=ev_gr.id"
    
    "WHERE res_gr_alumnos.id='%s'"
        "AND res_gr_alumnos.resourcetype='Class'"
        "AND res_aulas.resourcetype='Room'"
        
        "ORDER BY events_cuatrimestre2_1718.time;"
)

query4=("""SELECT events_cuatrimestre2_1718.eventgroup, events_cuatrimestre2_1718.id, events_cuatrimestre2_1718.time, res_aulas.name, res_aulas.color, res_gr_alumnos.id, ev_gr.name FROM  events_cuatrimestre2_1718
    INNER JOIN events_need_resources_c2_1718 enr_aulas ON enr_aulas.event= events_cuatrimestre2_1718.id
    INNER JOIN resources res_aulas ON enr_aulas.resource=res_aulas.id
    INNER JOIN events_need_resources_c2_1718 enr_gr_alumnos ON enr_gr_alumnos.event= events_cuatrimestre2_1718.id
    INNER JOIN resources res_gr_alumnos ON enr_gr_alumnos.resource=res_gr_alumnos.id
    INNER JOIN eventgroups ev_gr ON events_cuatrimestre2_1718.eventgroup=ev_gr.id
    
    WHERE res_aulas.resourcetype='Room'
        AND res_gr_alumnos.resourcetype='Class'
        AND res_gr_alumnos.id=
        """)


#Acceso al archivo xlsx (se crea y sobreescribe uno nuevo cada vez)
wb = openpyxl.Workbook()
 
wb.save(filepath)


#se repite la ejecución por cada par (grupo de alumnos,hoja de excel)
for (grupo, hoja) in pares_grupo_hoja:
    
    #Se extraen los eventos de un grupo
    cursor.execute(query4 + grupo)
    
    dictHorario = collections.defaultdict(lambda: [])
    
    #se almacenan los eventos en un diccionario de acuerdo al día de la semana correspondiente
    i=0
    for row in cursor:
        dictHorario[row[2]].append(row)
        print row[1]+"\n"+row[6]+"\n"+row[3]
        i=i+1
            
    print i
        
    # for k, v in dictHorario.iteritems():
    #     print k,v
     
    #Crear y seleccionar la hoja correspondiente
    wb.create_sheet(hoja)
    
        
    sheet1 = wb[hoja]
    sheet1["A1"]="#"
    wb.save(filepath)
    
    #Se recorre el diccionario iterando según claves y valores y se introduce la información en la tabla de excel
    for time_slot, cell_index in days_times.iteritems():
        print time_slot, cell_index
        #Si hay algún evento ese día se introduce
        if len(dictHorario[time_slot]) > 0:
            row=dictHorario[time_slot][0]
    #         print row[1]+" - "+row[6]+" - "+row[3]
    #         insert_value=row[1]+" - "+row[6]+" - "+row[3]
    #         print insert_value
    #         unaccented=unidecode.unidecode(insert_value)
    #         print unaccented+' '+str(type(unaccented))
    #         #u' '.join((agent_contact, agent_telno)).encode('utf-8').strip()
    #         sheet1[cell_index].value = unaccented
            sheet1[cell_index]=row[1]+" - "+row[6]+" - "+row[3]
            #CHANGE CELL COLOR #sheet1[cell_index].
            my_red = openpyxl.styles.colors.Color(rgb=row[4])
            my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
            sheet1[cell_index].fill = my_fill
        #Los eventos aparte del primero se introducen precedidos de un ;  
        if len(dictHorario[time_slot])>1:
            for i in range(1,len(dictHorario[time_slot])):
                row=dictHorario[time_slot][i]
                sheet1[cell_index]=sheet1[cell_index].value+"; "+row[1]+" - "+row[6]+" - "+row[3]
    
    
    
    #Sobreescribimos el archivo 
    wb.save(filepath)






####QUERIES
'''
SELECT events_cuatrimestre2_1718.eventgroup, events_cuatrimestre2_1718.id, events_cuatrimestre2_1718.time, res_aulas.name, res_grAlumnos.name FROM  events_cuatrimestre2_1718
    INNER JOIN events_need_resources_c2_1718 enr_aulas ON enr_aulas.event= events_cuatrimestre2_1718.id
    INNER JOIN resources res_aulas ON enr_aulas.resource=res_aulas.id
    INNER JOIN events_need_resources_c2_1718 enr_grAlumnos ON enr_grAlumnos.event= events_cuatrimestre2_1718.id
    INNER JOIN resources res_grAlumnos ON enr_grAlumnos.resource=res_grAlumnos.id
    
    WHERE events_cuatrimestre2_1718.eventgroup='CODIGO_ASIGNATURA_AQUI' 
        AND res_grAlumnos.resourcetype='Class'
        AND res_aulas.resourcetype='Room';
        
        
"SELECT events_cuatrimestre2_1718.eventgroup, events_cuatrimestre2_1718.id, events_cuatrimestre2_1718.time, res_aulas.name, res_grAlumnos.name FROM  events_cuatrimestre2_1718"
    "INNER JOIN events_need_resources_c2_1718 enr_aulas ON enr_aulas.event= events_cuatrimestre2_1718.id"
    "INNER JOIN resources res_aulas ON enr_aulas.resource=res_aulas.id"
    "INNER JOIN events_need_resources_c2_1718 enr_grAlumnos ON enr_grAlumnos.event= events_cuatrimestre2_1718.id"
    "INNER JOIN resources res_grAlumnos ON enr_grAlumnos.resource=res_grAlumnos.id"
    
    "WHERE events_cuatrimestre2_1718.eventgroup='%s'"   #introducir aquí código de asignatura 
        "AND res_grAlumnos.resourcetype='Class'"
        "AND res_aulas.resourcetype='Room';"
'''


'''
USE scheduledb;
SELECT events_cuatrimestre2_1718.eventgroup, events_cuatrimestre2_1718.id, events_cuatrimestre2_1718.time, res_aulas.name, res_aulas.color, res_gr_alumnos.id FROM  events_cuatrimestre2_1718
    INNER JOIN events_need_resources_c2_1718 enr_aulas ON enr_aulas.event= events_cuatrimestre2_1718.id
    INNER JOIN resources res_aulas ON enr_aulas.resource=res_aulas.id
    INNER JOIN events_need_resources_c2_1718 enr_gr_alumnos ON enr_gr_alumnos.event= events_cuatrimestre2_1718.id
    INNER JOIN resources res_gr_alumnos ON enr_gr_alumnos.resource=res_gr_alumnos.id
    
    WHERE #events_cuatrimestre2_1718.eventgroup='G100' AND 
        res_gr_alumnos.id='Class1_INFORMATICA_PLACEHOLDER'
        AND res_gr_alumnos.resourcetype='Class'
        AND res_aulas.resourcetype='Room'
        
        ORDER BY events_cuatrimestre2_1718.time
        ;
'''

'''
SELECT events_cuatrimestre2_1718.eventgroup, events_cuatrimestre2_1718.id, events_cuatrimestre2_1718.time, res_aulas.name, res_aulas.color, res_gr_alumnos.id, ev_gr.name FROM  events_cuatrimestre2_1718
    INNER JOIN events_need_resources_c2_1718 enr_aulas ON enr_aulas.event= events_cuatrimestre2_1718.id
    INNER JOIN resources res_aulas ON enr_aulas.resource=res_aulas.id
    INNER JOIN events_need_resources_c2_1718 enr_gr_alumnos ON enr_gr_alumnos.event= events_cuatrimestre2_1718.id
    INNER JOIN resources res_gr_alumnos ON enr_gr_alumnos.resource=res_gr_alumnos.id
    INNER JOIN eventgroups ev_gr ON events_cuatrimestre2_1718.eventgroup=ev_gr.id
    
    WHERE res_gr_alumnos.id=%s
        AND res_gr_alumnos.resourcetype='Class'
        AND res_aulas.resourcetype='Room'
        
        ORDER BY events_cuatrimestre2_1718.time
        ;
'''

